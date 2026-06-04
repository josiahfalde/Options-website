#!/usr/bin/env python3
"""
export_workbook.py  --  Wheel workbook -> normalized JSON bridge.

Converts "Options Trading.xlsx" (the per-company Wheel tracker) into the
normalized trade list the website consumes. This is the robust, offline bridge:
the browser can also import the .xlsx directly (src/data/importXlsx.ts uses the
same column layout), but this script lets us bake a real-data seed into the app
and re-export any time the workbook changes.

Usage:
    python tools/export_workbook.py \
        --src "C:/Users/josia/OneDrive/Documents/Options Trading.xlsx" \
        --out src/data/seed.json

The workbook is read-only here (openpyxl). If the file is locked by Excel/OneDrive
we copy it to a temp file first and read the copy -- writes never touch the
original (that is the COM-driven optionstrade skill's job).
"""
from __future__ import annotations
import argparse, json, os, shutil, tempfile, sys
from datetime import datetime, date

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required:  pip install openpyxl")

# ---- Layout of each "<TICKER> Wheel" sheet (0-indexed rows/cols) ------------
# row 0 : A1 live per-share price (#VALUE! until linked); col E (idx4) = value of 100 sh
# row 2 : ticker symbol in col A
# row 3 : block titles
# row 4 : column headers
# row 5+: data
# Credits block  cols 0..5 : Action, Date, Shares, Premium, Total, Status
# Mini summary   cols 6..7 : Credit / Debit-Assign / P&L / Net Cost Basis / Shares owned
# Debits block   cols 9..13: Action, Date, Shares, Cost, Invested
CREDIT_ACTION, CREDIT_DATE, CREDIT_SHARES, CREDIT_PREM, CREDIT_TOTAL, CREDIT_STATUS = 0, 1, 2, 3, 4, 5
DEBIT_ACTION, DEBIT_DATE, DEBIT_SHARES, DEBIT_COST, DEBIT_INVESTED = 9, 10, 11, 12, 13
SHARE_VALUE_COL = 4  # col E on row 0 holds value of 100 shares
FIRST_DATA_ROW = 5


def _num(v):
    if isinstance(v, (int, float)):
        return float(v)
    return None


def _date(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return None


def parse_sheet(ws, ticker):
    rows = list(ws.iter_rows(values_only=True))

    def cell(r, c):
        if r < len(rows) and c < len(rows[r]):
            return rows[r][c]
        return None

    # last-known price of 100 shares (col E, row 0) -> per share
    sh100 = _num(cell(0, SHARE_VALUE_COL))
    last_price = round(sh100 / 100.0, 4) if sh100 else None

    trades = []
    for r in range(FIRST_DATA_ROW, len(rows)):
        # ---- credit (sold contract) ----
        action = cell(r, CREDIT_ACTION)
        if isinstance(action, str) and action.strip():
            total = _num(cell(r, CREDIT_TOTAL))
            prem = _num(cell(r, CREDIT_PREM))
            amount = total if total is not None else prem
            d = _date(cell(r, CREDIT_DATE))
            if d and amount is not None:
                status = cell(r, CREDIT_STATUS)
                trades.append({
                    "ticker": ticker,
                    "side": "credit",
                    "action": action.strip(),           # CSP | CC
                    "date": d,
                    "shares": int(_num(cell(r, CREDIT_SHARES)) or 100),
                    "amount": round(amount, 2),
                    "status": (status.strip() if isinstance(status, str) else "Open"),
                })
        # ---- debit (buyback / assignment) ----
        daction = cell(r, DEBIT_ACTION)
        if isinstance(daction, str) and daction.strip():
            d = _date(cell(r, DEBIT_DATE))
            cost = _num(cell(r, DEBIT_COST))
            invested = _num(cell(r, DEBIT_INVESTED))
            if d:
                trades.append({
                    "ticker": ticker,
                    "side": "debit",
                    "action": daction.strip(),          # BB | AAssignSTK
                    "date": d,
                    "shares": int(_num(cell(r, DEBIT_SHARES)) or 100),
                    "amount": round(cost, 2) if cost is not None else 0.0,
                    "invested": round(invested, 2) if invested is not None else None,
                    "status": "",
                })
    return trades, last_price


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", default=r"C:/Users/josia/OneDrive/Documents/Options Trading.xlsx")
    # writes to a gitignored path by default so your REAL data never lands in
    # the public repo. Import the resulting JSON on the Data page (browser-only).
    ap.add_argument("--out", default="private/seed.real.json")
    args = ap.parse_args()

    src = args.src
    tmp = None
    try:
        wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    except PermissionError:
        tmp = os.path.join(tempfile.gettempdir(), "ot_export_copy.xlsx")
        # shared-read copy so an open/syncing workbook can still be read
        with open(src, "rb") as f_in, open(tmp, "wb") as f_out:
            shutil.copyfileobj(f_in, f_out)
        wb = openpyxl.load_workbook(tmp, read_only=True, data_only=True)

    all_trades = []
    prices = {}
    for name in wb.sheetnames:
        if not name.endswith("Wheel"):
            continue
        ticker = name.replace("Wheel", "").strip().upper()
        if not ticker:
            continue
        ws = wb[name]
        trades, last_price = parse_sheet(ws, ticker)
        all_trades.extend(trades)
        if last_price:
            prices[ticker] = last_price

    all_trades.sort(key=lambda t: (t["date"], t["ticker"]))

    # account-level config pulled from the Summary sheet
    capital_base = 2829.81
    spy = {"now": 754.13, "year_ago": 587.46, "year_ago_date": "2025-05-16"}
    if "Summary" in wb.sheetnames:
        srows = list(wb["Summary"].iter_rows(values_only=True))
        for row in srows:
            if not row:
                continue
            label = row[0] if isinstance(row[0], str) else ""
            if label == "Account capital base" and _num(row[1]):
                capital_base = _num(row[1])
            if label == "SPY price now (live)" and _num(row[1]):
                spy["now"] = _num(row[1])
            if label == "SPY one year ago" and _num(row[1]):
                spy["year_ago"] = _num(row[1])

    out = {
        "exportedAt": datetime.now().strftime("%Y-%m-%d"),
        "capitalBase": round(capital_base, 2),
        "lastPrices": prices,
        "spy": spy,
        "trades": all_trades,
    }

    os.makedirs(os.path.dirname(args.out) or ".", exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(out, f, indent=2)

    if tmp and os.path.exists(tmp):
        os.remove(tmp)

    print(f"Wrote {args.out}: {len(all_trades)} trades, "
          f"{len(prices)} tickers, capital base ${capital_base:,.2f}")


if __name__ == "__main__":
    main()
